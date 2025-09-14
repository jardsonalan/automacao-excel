import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def format_excel(path: str, sheet_name: str = 'Tabela'):
  wb = openpyxl.load_workbook(path)
  ws = wb[sheet_name]

  # Cor de fundo do cabeçalho
  header_fill = PatternFill(start_color='999999', end_color='999999', fill_type='solid')
  # Formatação da fonte do cabeçalho
  header_font = Font(name='Arial', size=11, bold=True, color='000000')
  # Cor de fundo dos dados
  data_fill = PatternFill(start_color='EEECE1', end_color='EEECE1', fill_type='solid')
  # Formatação da fonte dos dados
  data_font = Font(name='Arial', size=11, bold=False, color='000000')

  # Cor de fundo dos dados númericos
  number_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

  thin_side = Side(style='thin')

  # Aplicar bordas externas e estilos
  for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row), start=1):
      for col_idx, cell in enumerate(row, start=1):
          # Aplica a formatação das fontes e alinhamento
          cell.font = header_font if row_idx == 1 else data_font
          cell.alignment = Alignment(horizontal='center')

          # Preenche a cor de fundo do cabeçalho
          if row_idx == 1:
              cell.fill = header_fill
          else:
              # Preenche a cor de fundo das colunas dos (Ôxidos) e dos (Elementos)
              if col_idx in [1, 3]:
                  cell.fill = data_fill
              else:
                  # Preenche a cor de fundo das colunas númericas
                  cell.fill = number_fill

          # Aplica as bordas apenas na parte externa da tabela
          left = thin_side if col_idx == 1 else None
          right = thin_side if col_idx == ws.max_column else None
          top = thin_side if row_idx == 1 else None
          bottom = thin_side if row_idx == ws.max_row else None

          # Adiciona uma borda centralizada na tabela
          if col_idx == 2:
              right = thin_side

          cell.border = Border(left=left, right=right, top=top, bottom=bottom)

          # Formata os dados com valores númericos para apenas duas casas decimais
          if isinstance(cell.value, (int, float)):
              cell.number_format = '0.00'

  # Ajusta a largura das colunas
  for col_letter in ['A', 'B', 'C', 'D']:
      ws.column_dimensions[col_letter].width = 15

  # Salva a formatação da planilha
  wb.save(path)